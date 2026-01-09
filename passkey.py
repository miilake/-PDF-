#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import sys
import shutil
from typing import Dict, Tuple, Optional

import openpyxl

try:
    import pikepdf
except ImportError:
    print("缺少依赖 pikepdf。请先运行：pip install pikepdf", file=sys.stderr)
    raise


HEADER_ALIASES = {
    "filename": {"filename", "file", "name", "pdf", "pdfname", "文件名", "文件", "名称"},
    "password": {"password", "pwd", "pass", "密码"},
}


def norm(s) -> str:
    return str(s).strip().lower()


def detect_columns(first_row_values) -> Tuple[Optional[int], Optional[int], bool]:
    """
    返回 (filename_col_idx, password_col_idx, has_header)
    col_idx 为 0-based
    """
    values = [norm(v) if v is not None else "" for v in first_row_values]

    fn_idx = None
    pw_idx = None

    for i, v in enumerate(values):
        if v in HEADER_ALIASES["filename"]:
            fn_idx = i
        if v in HEADER_ALIASES["password"]:
            pw_idx = i

    has_header = (fn_idx is not None) or (pw_idx is not None)

    if not has_header:
        # 没有表头时，默认前两列
        return 0, 1, False

    # 有表头但缺列时，尽量兜底
    if fn_idx is None:
        fn_idx = 0
    if pw_idx is None:
        pw_idx = 1
    return fn_idx, pw_idx, True


def load_mapping_from_excel(excel_path: str, sheet: Optional[str] = None) -> Dict[str, str]:
    """
    读取 Excel，返回 {filename: password} 映射
    filename 允许不带 .pdf，会自动补全匹配
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"找不到工作表：{sheet}，现有工作表：{wb.sheetnames}")
        ws = wb[sheet]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 工作表为空")

    fn_col, pw_col, has_header = detect_columns(rows[0])

    start_idx = 1 if has_header else 0

    mapping: Dict[str, str] = {}
    for r in rows[start_idx:]:
        if r is None:
            continue
        fn = r[fn_col] if fn_col < len(r) else None
        pw = r[pw_col] if pw_col < len(r) else None

        if fn is None or str(fn).strip() == "":
            continue
        if pw is None or str(pw).strip() == "":
            # 没密码就跳过（也可以改成允许空密码）
            continue

        filename = str(fn).strip()
        password = str(pw).strip()

        mapping[filename] = password

    return mapping


def resolve_pdf_path(cwd: str, filename_from_excel: str) -> Optional[str]:
    """
    在 cwd 下解析 PDF 路径：
    - 支持 Excel 中写 "a.pdf" 或 "a"
    - 支持大小写不同的扩展名
    """
    cand = filename_from_excel.strip()

    # 若已经带 .pdf
    if cand.lower().endswith(".pdf"):
        path = os.path.join(cwd, cand)
        if os.path.isfile(path):
            return path
        # 尝试大小写不敏感匹配
        base = cand
    else:
        base = cand + ".pdf"

    # 直接匹配
    direct = os.path.join(cwd, base)
    if os.path.isfile(direct):
        return direct

    # 在目录里做一次不区分大小写的文件名匹配
    target_lower = base.lower()
    for f in os.listdir(cwd):
        if f.lower() == target_lower and os.path.isfile(os.path.join(cwd, f)):
            return os.path.join(cwd, f)

    return None


def encrypt_pdf(in_path: str, out_path: str, user_password: str, owner_password: Optional[str] = None) -> None:
    """
    使用 pikepdf 对 PDF 加密。默认 owner_password = user_password
    使用 R=6（AES-256）尽可能提高强度。
    """
    owner_pw = owner_password if owner_password is not None else user_password

    # 如果已加密且不知道原密码，Pdf.open 会报错
    with pikepdf.Pdf.open(in_path) as pdf:
        pdf.save(
            out_path,
            encryption=pikepdf.Encryption(
                user=user_password,
                owner=owner_pw,
                R=6,  # AES-256（若 PDF 版本允许）
            ),
        )


def safe_replace(src_tmp: str, dst_final: str) -> None:
    """
    覆盖写入时，先写临时文件，再原子替换。
    """
    # Windows 下 os.replace 也可用
    os.replace(src_tmp, dst_final)


def main():
    parser = argparse.ArgumentParser(
        description="从 Excel 读取 (文件名, 密码)，批量给当前文件夹下的 PDF 加密码。默认输出到 protected/。",
    )
    parser.add_argument("--excel", required=True, help="Excel 文件路径，如：passwords.xlsx")
    parser.add_argument("--sheet", default=None, help="工作表名称（可选），默认使用第一个激活工作表")
    parser.add_argument("--outdir", default="protected", help="输出目录（默认 protected）。当 --inplace 时忽略")
    parser.add_argument("--inplace", action="store_true", help="直接覆盖原 PDF（慎用，建议先备份）")
    parser.add_argument("--owner", default=None, help="所有者密码（可选）。不填则与用户密码相同")
    args = parser.parse_args()

    cwd = os.getcwd()

    if not os.path.isfile(args.excel):
        print(f"[ERROR] 找不到 Excel：{args.excel}", file=sys.stderr)
        sys.exit(2)

    try:
        mapping = load_mapping_from_excel(args.excel, sheet=args.sheet)
    except Exception as e:
        print(f"[ERROR] 读取 Excel 失败：{e}", file=sys.stderr)
        sys.exit(2)

    if not mapping:
        print("[ERROR] Excel 中未读取到任何有效 (文件名, 密码) 记录。", file=sys.stderr)
        sys.exit(2)

    if not args.inplace:
        outdir = os.path.join(cwd, args.outdir)
        os.makedirs(outdir, exist_ok=True)

    total = 0
    ok = 0
    skipped = 0
    failed = 0

    for excel_name, password in mapping.items():
        total += 1
        in_pdf = resolve_pdf_path(cwd, excel_name)
        if not in_pdf:
            print(f"[SKIP] 未找到对应 PDF：{excel_name}")
            skipped += 1
            continue

        try:
            if args.inplace:
                tmp_path = in_pdf + ".tmp_encrypt"
                encrypt_pdf(in_pdf, tmp_path, password, owner_password=args.owner)
                safe_replace(tmp_path, in_pdf)
                print(f"[OK] 已覆盖加密：{os.path.basename(in_pdf)}")
            else:
                out_path = os.path.join(outdir, os.path.basename(in_pdf))
                encrypt_pdf(in_pdf, out_path, password, owner_password=args.owner)
                print(f"[OK] 已输出加密：{os.path.basename(in_pdf)} -> {args.outdir}/")
            ok += 1

        except pikepdf._qpdf.PasswordError:
            print(f"[FAIL] PDF 已加密且无法打开（需要原密码）：{os.path.basename(in_pdf)}", file=sys.stderr)
            failed += 1
        except Exception as e:
            print(f"[FAIL] 加密失败：{os.path.basename(in_pdf)}，原因：{e}", file=sys.stderr)
            failed += 1

    print("\n==== 处理结果 ====")
    print(f"总记录：{total}")
    print(f"成功：{ok}")
    print(f"跳过（找不到文件等）：{skipped}")
    print(f"失败：{failed}")

    # 若全部失败，返回非 0 方便脚本化
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
